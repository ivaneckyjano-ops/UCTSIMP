from __future__ import annotations

import sqlite3
from decimal import Decimal
from pathlib import Path

from .fifo_realized import fifo_danove_z_db
from .models import CashflowSummary, DailyNetRow, SummaryRow, TaxSplitCashflow
from .tax_relevance import sql_nedanove_kategorie_in

GROSS_EUR_VYSVETLENIE = (
    "Stlpec **Gross EUR** v prehlade pod tickermi (a v suhrne pod kategoriou `trade`) je sucet stlpca "
    "„Gross Amount“ z IBKR pre riadky Buy/Sell, prevedeny do **EUR** — ide o **ciastkove (hrube) nohy obchodu**, "
    "nie o **zrealizovany vysledok (realized P&L)** z uzavretych pozicii. Ten by vyzadoval sparovanie nákupu "
    "a predaja (FIFO, opcie, expirácia atd.).\n\n"
    "Na penazny dopad a cashflow sa pozerajte na stlpec **Net EUR** (netto po provizii) a na zalozku **Prehlad** "
    "— sučet príjmov, výdajov a bežiaci kumulativ z importu.\n\n"
    "Riadok **danovy (FIFO)**: u obchodov (`trade`) sa nepouziva cisty Net z kazdeho riadka; pocita sa "
    "**realizovany vysledok** po symboli (FIFO) — teda az pri uzatvaracich nákup+predaj. Otvorene nohy: 0. "
    "Vklady a vybery: neda. Ostatne kategorie: Net EUR. Over s danovym poradcom."
)


def ticker_summary(connection: sqlite3.Connection) -> list[SummaryRow]:
    return _summary(
        connection,
        """
        SELECT COALESCE(ticker, 'Bez tickera') AS key,
               COUNT(*) AS trade_count,
               SUM(CAST(gross_amount_eur AS REAL)) AS gross_eur,
               SUM(CAST(commission_eur AS REAL)) AS commission_eur,
               SUM(CAST(net_amount_eur AS REAL)) AS net_eur
        FROM transactions
        WHERE category = 'trade'
        GROUP BY COALESCE(ticker, 'Bez tickera')
        ORDER BY key
        """,
    )


def category_summary(connection: sqlite3.Connection) -> list[SummaryRow]:
    return _summary(
        connection,
        """
        SELECT category AS key,
               COUNT(*) AS trade_count,
               SUM(CAST(gross_amount_eur AS REAL)) AS gross_eur,
               SUM(CAST(commission_eur AS REAL)) AS commission_eur,
               SUM(CAST(net_amount_eur AS REAL)) AS net_eur
        FROM transactions
        GROUP BY category
        ORDER BY category
        """,
    )


def cashflow_summary(connection: sqlite3.Connection) -> CashflowSummary:
    row = connection.execute(
        """
        SELECT
            COALESCE(
                SUM(
                    CASE
                        WHEN CAST(net_amount_eur AS REAL) > 0
                        THEN CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS prijem,
            COALESCE(
                SUM(
                    CASE
                        WHEN CAST(net_amount_eur AS REAL) < 0
                        THEN -CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS vydaj,
            COALESCE(SUM(CAST(net_amount_eur AS REAL)), 0.0) AS cisty
        FROM transactions
        """
    ).fetchone()
    assert row is not None
    return CashflowSummary(
        prijem_eur=_decimal_from_sql(row["prijem"]),
        vydaj_eur=_decimal_from_sql(row["vydaj"]),
        cisty_pohyb_eur=_decimal_from_sql(row["cisty"]),
    )


def tax_split_cashflow(connection: sqlite3.Connection) -> TaxSplitCashflow:
    """
    Neda: vklady/vybery. Dan netrade: provizia, uroky, … podla Net EUR.
    Obchody (`trade`): **iba realizovany** vysledok (FIFO) — parovane/uzatvaraci predaj proti naku
    otvorene nohy: 0 v dan. obch. casti.
    """
    in_list = sql_nedanove_kategorie_in()
    row = connection.execute(
        f"""
        SELECT
            COALESCE(
                SUM(
                    CASE
                        WHEN CAST(net_amount_eur AS REAL) > 0
                             AND category NOT IN ({in_list})
                             AND category != 'trade'
                        THEN CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS prijem_dan_net,
            COALESCE(
                SUM(
                    CASE
                        WHEN CAST(net_amount_eur AS REAL) > 0
                             AND category IN ({in_list})
                        THEN CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS prijem_nedan,
            COALESCE(
                SUM(
                    CASE
                        WHEN CAST(net_amount_eur AS REAL) < 0
                             AND category NOT IN ({in_list})
                             AND category != 'trade'
                        THEN -CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS vydaj_dan_net,
            COALESCE(
                SUM(
                    CASE
                        WHEN CAST(net_amount_eur AS REAL) < 0
                             AND category IN ({in_list})
                        THEN -CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS vydaj_nedan,
            COALESCE(
                SUM(
                    CASE
                        WHEN category NOT IN ({in_list})
                         AND category != 'trade'
                        THEN CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS cisty_dan_net,
            COALESCE(
                SUM(
                    CASE
                        WHEN category IN ({in_list})
                        THEN CAST(net_amount_eur AS REAL)
                        ELSE 0.0
                    END
                ),
                0.0
            ) AS cisty_nedan
        FROM transactions
        """
    ).fetchone()
    assert row is not None
    t_pr, t_vd, t_c = fifo_danove_z_db(connection)
    p_dan = _decimal_from_sql(row["prijem_dan_net"]) + t_pr
    v_dan = _decimal_from_sql(row["vydaj_dan_net"]) + t_vd
    c_dan = _decimal_from_sql(row["cisty_dan_net"]) + t_c
    return TaxSplitCashflow(
        prijem_danovy_eur=p_dan,
        prijem_nedanovy_eur=_decimal_from_sql(row["prijem_nedan"]),
        vydaj_danovy_eur=v_dan,
        vydaj_nedanovy_eur=_decimal_from_sql(row["vydaj_nedan"]),
        cisty_danovy_eur=c_dan,
        cisty_nedanovy_eur=_decimal_from_sql(row["cisty_nedan"]),
    )


def daily_cumulative_net(connection: sqlite3.Connection) -> list[DailyNetRow]:
    """Denné súčty Net EUR a kumulatív od prvého dátumu v (importe) databáze."""
    cur = connection.execute(
        """
        WITH den AS (
            SELECT
                trade_date,
                SUM(CAST(net_amount_eur AS REAL)) AS denna_zmena
            FROM transactions
            GROUP BY trade_date
        )
        SELECT
            trade_date,
            denna_zmena,
            SUM(denna_zmena) OVER (
                ORDER BY trade_date
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            ) AS kumulativ
        FROM den
        ORDER BY trade_date
        """
    )
    return [
        DailyNetRow(
            obchodny_den=str(r["trade_date"]),
            denna_zmena_eur=_decimal_from_sql(r["denna_zmena"]),
            kumulativ_eur=_decimal_from_sql(r["kumulativ"]),
        )
        for r in cur.fetchall()
    ]


def yearly_summary(connection: sqlite3.Connection) -> list[SummaryRow]:
    return _summary(
        connection,
        """
        SELECT substr(trade_date, 1, 4) AS key,
               COUNT(*) AS trade_count,
               SUM(CAST(gross_amount_eur AS REAL)) AS gross_eur,
               SUM(CAST(commission_eur AS REAL)) AS commission_eur,
               SUM(CAST(net_amount_eur AS REAL)) AS net_eur
        FROM transactions
        GROUP BY substr(trade_date, 1, 4)
        ORDER BY key
        """,
    )


def review_rows(connection: sqlite3.Connection) -> list[sqlite3.Row]:
    return list(
        connection.execute(
            """
            SELECT trade_date, description, symbol, transaction_type, category,
                   net_amount_eur
            FROM transactions
            WHERE needs_review = 1
            ORDER BY trade_date, id
            """
        )
    )


def export_excel(connection: sqlite3.Connection, path: str | Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Suhrn"
    _write_summary_block(summary_sheet, "Podla tickerov", ticker_summary(connection), 1)
    _write_summary_block(summary_sheet, "Podla kategorii", category_summary(connection), 1 + len(ticker_summary(connection)) + 4)
    _write_summary_block(summary_sheet, "Podla rokov", yearly_summary(connection), 1 + len(ticker_summary(connection)) + len(category_summary(connection)) + 8)

    flow = cashflow_summary(connection)
    tax = tax_split_cashflow(connection)
    flow_sheet = workbook.create_sheet("PrijemVydaj")
    flow_sheet.append(["Polozka", "EUR"])
    flow_sheet.append(["Celkovo: Prijem (sucet kladnych Net EUR)", float(flow.prijem_eur)])
    flow_sheet.append(["Celkovo: Vydaj (sucet |zapornych| Net EUR)", float(flow.vydaj_eur)])
    flow_sheet.append(["Celkovo: Cisty pohyb (sucet vsetkych Net EUR)", float(flow.cisty_pohyb_eur)])
    flow_sheet.append([])
    flow_sheet.append(
        ["Danovy: Prijem (ostatne Net + FIFO realiz. obchodov)", float(tax.prijem_danovy_eur)]
    )
    flow_sheet.append(
        ["Danovy: Vydaj (ostatne + |FIFO zapor. realiz.|)", float(tax.vydaj_danovy_eur)]
    )
    flow_sheet.append(
        ["Danovy: Cisty (ostatne Net + suhrn FIFO obchodov)", float(tax.cisty_danovy_eur)]
    )
    flow_sheet.append([])
    flow_sheet.append(["Nedanovy tok: Prijem (kl - vklady...)", float(tax.prijem_nedanovy_eur)])
    flow_sheet.append(["Nedanovy tok: Vydaj (|zap.| vybery...)", float(tax.vydaj_nedanovy_eur)])
    flow_sheet.append(["Nedanovy tok: Cisty (sucet Net v nedan. kateg.)", float(tax.cisty_nedanovy_eur)])
    flow_sheet.append([])
    flow_sheet.append(
        [
            "Poznamka k Gross EUR a dan/ neda",
            GROSS_EUR_VYSVETLENIE.replace("\n\n", " ").replace("**", ""),
        ]
    )

    kum_sheet = workbook.create_sheet("Kumulativ")
    kum_sheet.append(["Datum", "Denna zmena Net EUR", "Kumulativ Net EUR"])
    for r in daily_cumulative_net(connection):
        kum_sheet.append(
            [r.obchodny_den, float(r.denna_zmena_eur), float(r.kumulativ_eur)]
        )

    transaction_sheet = workbook.create_sheet("Transakcie")
    transaction_headers = [
        "Datum",
        "Ticker",
        "Symbol",
        "Popis",
        "Typ",
        "Kategoria",
        "Mnozstvo",
        "Cena",
        "Mena ceny",
        "Gross EUR",
        "Komisia EUR",
        "Net EUR",
        "Na kontrolu",
    ]
    transaction_sheet.append(transaction_headers)
    for row in connection.execute(
        """
        SELECT trade_date, ticker, symbol, description, transaction_type, category,
               quantity, price, price_currency, gross_amount_eur, commission_eur,
               net_amount_eur, needs_review
        FROM transactions
        ORDER BY trade_date, id
        """
    ):
        transaction_sheet.append(
            [
                row["trade_date"],
                row["ticker"],
                row["symbol"],
                row["description"],
                row["transaction_type"],
                row["category"],
                _number_or_text(row["quantity"]),
                _number_or_text(row["price"]),
                row["price_currency"],
                _number_or_text(row["gross_amount_eur"]),
                _number_or_text(row["commission_eur"]),
                _number_or_text(row["net_amount_eur"]),
                "ano" if row["needs_review"] else "",
            ]
        )

    review_sheet = workbook.create_sheet("Na kontrolu")
    review_sheet.append(["Datum", "Popis", "Symbol", "Typ", "Kategoria", "Net EUR"])
    for row in review_rows(connection):
        review_sheet.append(
            [
                row["trade_date"],
                row["description"],
                row["symbol"],
                row["transaction_type"],
                row["category"],
                _number_or_text(row["net_amount_eur"]),
            ]
        )

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)

    workbook.save(path)


def _summary(connection: sqlite3.Connection, sql: str) -> list[SummaryRow]:
    rows = connection.execute(sql).fetchall()
    return [
        SummaryRow(
            key=row["key"],
            trade_count=int(row["trade_count"]),
            gross_eur=_decimal_from_sql(row["gross_eur"]),
            commission_eur=_decimal_from_sql(row["commission_eur"]),
            net_eur=_decimal_from_sql(row["net_eur"]),
        )
        for row in rows
    ]


def _write_summary_block(
    sheet, title: str, rows: list[SummaryRow], start_row: int
) -> None:
    sheet.cell(row=start_row, column=1, value=title)
    headers = ["Kluc", "Pocet", "Gross EUR", "Komisia EUR", "Net EUR"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=start_row + 1, column=index, value=header)
    for row_index, row in enumerate(rows, start=start_row + 2):
        sheet.cell(row=row_index, column=1, value=row.key)
        sheet.cell(row=row_index, column=2, value=row.trade_count)
        sheet.cell(row=row_index, column=3, value=float(row.gross_eur))
        sheet.cell(row=row_index, column=4, value=float(row.commission_eur))
        sheet.cell(row=row_index, column=5, value=float(row.net_eur))


def _decimal_from_sql(value) -> Decimal:
    if value is None:
        return Decimal("0")
    return Decimal(str(value)).quantize(Decimal("0.01"))


def _number_or_text(value: str | None):
    if value is None:
        return None
    try:
        return float(value)
    except ValueError:
        return value
